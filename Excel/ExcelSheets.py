from openpyxl import Workbook

#==========
def print_rows():
    print ("========================")
    for row in mySheet.iter_rows(values_only=True):
        print(row)
    print ("========================")
#==========

fileFullPath = "/Users/yishi/Documents/GitHub/PythonLearning/Excel/Cars.xlsx"

myWorkbook = Workbook()
mySheet = myWorkbook.active

# Change sheet name
mySheet.title = "Overview"

mySheet["A1"] = "Brand"
mySheet["B1"] = "Model"
mySheet["C1"] = "Price"
print_rows()

# Create a new sheet
mySheet_2 = myWorkbook.create_sheet("Details")
print (myWorkbook.sheetnames)
# Create a new sheet as the first sheet
mySheet_0 = myWorkbook.create_sheet("First",0)
print (myWorkbook.sheetnames)

# Delete a sheet, need to pass the sheet object
myWorkbook.remove(mySheet_0)
print(myWorkbook.sheetnames)

# Copy a sheet
myWorkbook.copy_worksheet(mySheet_2)
print(myWorkbook.sheetnames)


myWorkbook.save(fileFullPath)

