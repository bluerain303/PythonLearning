from openpyxl import load_workbook

print ("=================")
# Open the file
wb = load_workbook("/Users/yishi/Documents/GitHub/PythonLearning/Excel/TreeData.xlsx")

# Get the active sheet
sheet = wb.active
#print (sheet.title)

# Get value of a cell
# print (sheet["A1"].value)
# Or
# print (sheet.cell(1,1).value)

# Get a range
# print (sheet["A1:C3"])

# Get all cells from column A (cells with values only)
# print (sheet["A"])

# Get all cells from a range of columns
# print (sheet["A:B"])

# Get all cells from row 2
# print (sheet[2])

# Get call cells from a range of rows
# print (sheet[1:3])

# Iteration - Row
'''
for row in sheet.iter_rows(min_row=1,max_row = 2, min_col=1, max_col=3):
    print (row)
'''

# Iteration - Row
'''
for col in sheet.iter_cols(min_row=1,max_row = 2, min_col=1, max_col=3):
    print (col)
'''

# Get values from iteration, use values_only
'''
for row in sheet.iter_rows(min_row=1,max_row = 2, min_col=1, max_col=3, values_only=True):
    print (row)
'''

# Go through ALL rows
for row in sheet.rows:
    print (row)


print ("=================")