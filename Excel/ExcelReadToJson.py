import json
from openpyxl import load_workbook

print ("=================")
# Open the file
wb = load_workbook("/Users/yishi/Documents/GitHub/PythonLearning/Excel/TreeData.xlsx")

# Get the active sheet
sheet = wb.active

trees = {}
for row in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3, values_only=True):
    treeName = row[0]
    treeValues = {
        "Color" : row[1],
        "Height": row[2]
    }
    trees[treeName] = treeValues

print (json.dumps(trees))

print ("=================")