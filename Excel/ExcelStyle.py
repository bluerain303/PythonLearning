from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

#==========
def print_rows():
    print ("========================")
    for row in mySheet.iter_rows(values_only=True):
        print(row)
    print ("========================")
#==========

fileFullPath = "/Users/yishi/Documents/GitHub/PythonLearning/Excel/Cars.xlsx"

myWorkbook = Workbook()
mySheet = myWorkbook.active

# Change sheet name
mySheet.title = "Overview"

# Set values
headeRow = ["Brand", "Model", "Prices"]
mySheet.append(headeRow)
BMW = ["BMW", "X1", 80]
MAZDA = ["MAZDA", "CX5", 40]
TOYOTA = ["TOYOTA", "RAV4", 50]
mySheet.append(BMW)
mySheet.append(MAZDA)
mySheet.append(TOYOTA)

print_rows()

# Conditional formatting.
redBackground = PatternFill(bgColor="00FF0000")
diff_style = DifferentialStyle(fill=redBackground)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$C2>65"]
mySheet.conditional_formatting.add("A2:D10", rule)
# Note: the formula should use the first line from the range i.e. 2.
# Otherwise, if use C1, then line 2 will be formatted according to line 1.




myWorkbook.save(fileFullPath)

